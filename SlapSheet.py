##bunch of imports
import json
import os
from datetime import datetime
import pandas as pd
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import warnings
#disable warning for deprecated functions, we don't CARE
warnings.simplefilter(action='ignore', category=FutureWarning)

## define some standard import stuff, we use this twice so saves a bit of space.
def import_json(filename):
    with open(filename, 'r', encoding='windows-1252', errors='ignore') as f:
        return json.load(f, strict=False)

### read settings
print('reading settings...')
with open('settings.cfg', 'r') as f:
    settings = f.readlines()
week = int(settings[0].split('=')[1].strip())
div = str(settings[1].split('=')[1].strip())
season = int(settings[2].split('=')[1].strip())
format = str(settings[3].split('=')[1].strip())

#based on the settings, chance div and change the sheet we're reading.
#note that the sheet name is hard coded. Not optimal but alas.
if div.lower() == 'pro':
    div = 'Pro'
    xlname = 'EUSL 4v4 Pro Division Statistics  (Rebound).xlsx'
if div.lower() == 'challenger':
    div = 'Challenger'
    xlname = 'EUSL 4v4 Challenger Division Statistics  (Rebound).xlsx'
if div.lower() == 'inter' or div.lower()=='intermediate':
    div = 'Intermediate'
    xlname = 'EUSL 4v4 Intermediate Division Statistics  (Rebound).xlsx'
if div.lower() == 'entry':
    div='Entry'
    xlname = 'EUSL 4v4 Entry Division Statistics  (Rebound).xlsx'

#read the stats sheet to extract players and matches
print('reading stats sheet...')
xls = pd.ExcelFile(xlname)

#convert players to pandas dataframe
players_from_sheet = pd.read_excel(xlname, sheet_name='Player_Teams')
#set the index
players_from_sheet.set_axis(['Season', 'Player', 'SlapID', 'Team', 'Query'], axis=1, inplace=True)
#convert matches to pandas dataframe
matches_from_sheet = pd.read_excel(xlname, sheet_name='Matches_Season')

#try to extract most recent match and extract match number, if start of season, set number to 0
try:
    most_recent_match = matches_from_sheet['Unnamed: 2'][1]
    most_recent_match_number = int(most_recent_match.split('G')[1].strip().split('C')[0].strip())
except:
    most_recent_match_number = 0

def recognize_team(log):
    ### recognize which team is playing
    home_team_counter = Counter() #counter objects are basically dictionaries that count
    away_team_counter = Counter()

    #loop over each player in log
    for player in log['players']:
        #check player team
        if player['team'] == 'home':
            try:
                #we always use the players name on the sheet, based on sheet SlapID matching game SlapID. Then we look up their team in sheet and add 1 to the count of that team.
                home_team_counter[players_from_sheet[players_from_sheet['SlapID'] == int(player['game_user_id'])]['Team'].values[0]] += 1
            except:
                #if the log's SLapID is not on the sheet, mark player as EA (could also be alt)
                home_team_counter['EA'] += 1
        else:
            try:
                away_team_counter[players_from_sheet[players_from_sheet['SlapID'] == int(player['game_user_id'])]['Team'].values[0]] += 1
            except:
                away_team_counter['EA'] += 1
    #now we extract the most common team counted from the players on the log. If that is EA, we use the second most common.
    #This should be a very rare edge case where a team uses more EA's than actual players.
    home_team = home_team_counter.most_common(2)[0][0] if not home_team_counter.most_common(2)[0][0] == 'EA' else home_team_counter.most_common(2)[1][0]
    away_team = away_team_counter.most_common(2)[0][0] if not away_team_counter.most_common(2)[0][0] == 'EA' else away_team_counter.most_common(2)[1][0]
    #finally return home team and away team
    return home_team, away_team


def log_to_match(log):
    '''This function transforms a match log into the appropriate pandas dataframe
        That we can later use to populate the excel sheet'''
    #initialise most recent match number as global variable - ugly
    global most_recent_match_number

    ### convert a match to a log
    if log['current_period'] == '3' or log['end_reason'] == 'MercyRule':
        #make a dict we can use to translate the attributes from the match log to the attributes as they are named on stats sheet
        json_to_excel = {'score':'Score', 'game_user_id':'Slap id', 'goals':'Goals', 'shots':'Shots', 'assists':'Assists', 'passes':'Passes', 'saves':'Saves', 'blocks':'Blocks', 'post_hits':'Post Hits', 'takeaways':'Takeaways', 'turnovers':'Turnovers', 'primary_assists':'Primary Assists', 'secondary_assists':'Secondary Assists','possession_time_sec':'Possession Time', 'faceoffs_won':'Faceoffs Won', 'faceoffs_lost':'Faceoffs Lost', 'wins':'Wins', 'losses':'Losses', 'ties':'Ties', 'overtime_losses':'OT Losses', 'overtime_wins':'OT Wins', 'game_winning_goals':'Game Winning Goals', 'shutouts': 'Shutout Periods', 'shutouts_against':'Shutout Periods Against', 'conceded_goals':'Goals Against', 'contributed_goals':'Goals For'}
        #initialize a dataframe to record match data in. We have to follow the exact amount of cells in the stats sheet
        #this makes it so there's a bunch of extra columns in here that do not actually get filled, like blank1, blank2
        #we need these extra columns to make the dataframe output fit the needed excel format. This is also ugly.
        match_df = pd.DataFrame(columns=['blank1', 'blank2', 'match_code','loss or win','Team','TScore','TGoals Against','Shots For','Shots Against', 'TSaves', 'Players','Slap id','Score','Goals','Assists','Shots','Saves','Passes','Blocks','Post Hits','Takeaways','Turnovers','Primary Assists','Secondary Assists','Possession Time','Faceoffs Won','Faceoffs Lost','Wins','Losses','Ties','OT Wins','OT Losses','Game Winning Goals','Shutout Periods','Shutout Periods Against','Goals For','Goals Against','Periods Played', 'some_shit', 'blank3', 'away_team_r', 'away_team_score', 'home_team_r', 'home_team_score'])

        #initialize some team stats by utilizing some clever list comprehension.
        #To extract these team stats we need to loop over every player, and sum their stats.
        #team stats are not actually recorded in the match logs themselves.
        shots_home = sum([player['stats']['shots'] if 'shots' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        shots_away = sum([player['stats']['shots'] if 'shots' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        saves_home = sum([player['stats']['saves'] if 'saves' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        saves_away = sum([player['stats']['saves'] if 'saves' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        goals_home = max([player['stats']['contributed_goals'] if 'contributed_goals' in player['stats'].keys() else 0 for player in log['players'] if
                          player['team'] == 'home'])
        goals_away = max([player['stats']['contributed_goals'] if 'contributed_goals' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        playercount_home = len([player for player in log['players'] if player['team'] == 'home'])
        playercount_away = len([player for player in log['players'] if player['team'] == 'away'])
        home_ties = sum([player['stats']['ties'] if 'ties' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        home_wins = sum([player['stats']['wins'] if 'wins' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        home_losses = sum([player['stats']['losses'] if 'losses' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        home_periods = home_ties + home_wins + home_losses
        home_ot_wins = sum([player['stats']['overtime_wins'] if 'overtime_wins' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        home_ot_losses = sum([player['stats']['overtime_losses'] if 'overtime_losses' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'home'])
        away_ties = sum([player['stats']['ties'] if 'ties' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        away_wins = sum([player['stats']['wins'] if 'wins' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        away_losses = sum([player['stats']['losses'] if 'losses' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        away_periods = away_ties + away_wins + away_losses
        away_ot_wins = sum([player['stats']['overtime_wins'] if 'overtime_wins' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])
        away_ot_losses = sum([player['stats']['overtime_losses'] if 'overtime_losses' in player['stats'].keys() else 0 for player in log['players'] if player['team'] == 'away'])

        #define the winner and get the names of the teams
        home_win = goals_home>goals_away
        winner = 'home' if home_win else 'away'
        home_team, away_team = recognize_team(log)

        #increment match number by one
        most_recent_match_number += 1
        #make the match code
        match_code = 'S{} W{} G{} C'.format(season, week, most_recent_match_number)

        #make a list of empty strings to fill our empty df with
        #we use this empty df to insert a blank line
        blank_list = [['' for i in range(len(match_df.columns))]]
        blank_df = pd.DataFrame(blank_list, columns=match_df.columns)

        #now make a df for the match code line
        #as you can see we use a bunch of dfs to get the right formatting in excel, its tedious AF
        match_code_list_blank = ['' for i in range(len(match_df.columns))]
        match_code_list_blank[2] = match_code
        match_code_blank_df = pd.DataFrame([match_code_list_blank], columns=match_df.columns)
        match_code_list = ['' for i in range(len(match_df.columns))]
        match_code_list[2] = match_code
        match_code_list[-5] = 'OT' if away_ot_losses > 0 or home_ot_losses > 0 else ''
        match_code_list[-4] = home_team
        match_code_list[-3] = f'{"L" if "away" == winner else "W"} {goals_home}'
        match_code_list[-2] = away_team
        match_code_list[-1] = f'{"L" if "home" == winner else "W"} {goals_away}'

        #now lets do a bunch of error checking, starting with correct periods
        if log['current_period'] == '3':
            if home_periods != 12:
                match_code_list[3] = 'ERROR: HOME TEAM PERIODS NOT 12, CHECK LOGS'
            if away_periods != 12:
                match_code_list[8] = 'ERROR: AWAY TEAM PERIODS NOT 12, CHECK LOGS'
        #the program only gets to this point if mercy is true,
        #so if the current period is 2:
        #it means it must have been a mercy and that means total periods should be periods*4
        elif log['current_period'] == '2':
            if home_periods != 8:
                match_code_list[3] = 'ERROR: MERCY IN 2 PERIODS BUT  HOME TEAM PERIODS NOT 8, CHECK LOGS'
            if away_periods != 8:
                match_code_list[8] = 'ERROR: MERCY IN 2 PERIODS BUT AWAY TEAM PERIODS NOT 8, CHECK LOGS'
        elif log['current_period'] == '1':
            if home_periods != 4:
                match_code_list[3] = 'ERROR: MERCY IN 1 PERIOD BUT HOME TEAM PERIODS NOT 4, CHECK LOGS'
            if away_periods != 4:
                match_code_list[8] = 'ERROR: MERCY IN 1 PERIOD BUT AWAY TEAM PERIODS NOT 4, CHECK LOGS'
        #we save potential errors on the match code line in different cells
        match_code_df = pd.DataFrame([match_code_list], columns=match_df.columns)
        match_df = match_df.append(blank_df)
        match_df = match_df.append(match_code_df)
        match_df = match_df.append(match_code_blank_df)

        #last step is to rename our columns.
        #in the excel doc we use column names like "Score" or "Saves" multiple times in the same row
        #pandas doesnt allow this, unless we use this dumb workaround of renaming them at the end.
        columnrenamer = {'blank1':'', 'blank2':'','TScore': 'Score', 'TShots': 'Shots', 'TSaves': 'Saves', 'loss or win': '', 'some_shit': '', 'TGoals Against': 'Goals Against', 'match_code':'', 'blank3':'', 'away_team_r':'', 'away_team_score':'', 'home_team_r':'', 'home_team_score':'', 'blank11':''}
        renamed_columns = [columnrenamer[column] if column in columnrenamer.keys() else column for column in match_df.columns]
        column_df = pd.DataFrame([[i for i in renamed_columns]], columns=match_df.columns)

        def team_checker(log, side, match_df):
            '''This function extracts all the players stats in a log for a certain side
            And returns a dataframe with players sorted by score.'''

            temp_df = pd.DataFrame()
            for player in log['players']:
                if player['team'] == side:
                    #check if player is on sheet by SlapID, and extract their username from sheet.
                    try:
                        playername_on_sheet = players_from_sheet[players_from_sheet['SlapID'] == int(player['game_user_id'])]['Player'].values[0]
                        player_slapid_on_sheet = players_from_sheet[players_from_sheet['SlapID'] == int(player['game_user_id'])]['SlapID'].values[0]
                    #if not on sheet, add - NOT IN SHEET to username.
                    except:
                        playername_on_sheet = player['username']+' - NOT IN SHEET'
                        player_slapid_on_sheet = player['game_user_id']

                    #initialize a dict of default values for every playerstat.
                    playerstats = {'Slap id': 0, 'Score': 0,
                                   'Goals': 0, 'Assists': 0, 'Shots': 0, 'Saves': 0, 'Passes': 0, 'Blocks': 0, 'Post Hits': 0,
                                   'Takeaways': 0, 'Turnovers': 0,
                                   'Primary Assists': 0, 'Secondary Assists': 0, 'Possession Time': 0, 'Faceoffs Won': 0,
                                   'Faceoffs Lost': 0, 'Wins': 0,
                                   'Losses': 0, 'Ties': 0, 'OT Wins': 0, 'OT Losses': 0, 'Game Winning Goals': 0,
                                   'Shutout Periods': 0,
                                   'Shutout Periods Against': 0, 'Goals For': 0, 'Goals Against': 0, 'Periods Played': 0}

                    for stat, val in player['stats'].items():
                        #for every stat, and it's according value
                        if stat in json_to_excel.keys(): #this checks if the stat is in the Excel sheet stats, safeguard for adding of new stats
                            playerstats[json_to_excel[stat]] = val
                            #add the value of the stat to the stat of the current player

                    #initialize a dataframe for the current player, with the right columns
                    player_df = pd.DataFrame([['', '', match_code, 'WIN' if side == winner else 'LOSS', home_team if side=='home' else away_team, goals_home if side == 'home' else goals_away,
                                                   goals_away if side=='home' else goals_home, shots_home if side == 'home' else shots_away,
                                                   shots_away if side=='home' else shots_home,
                                                   saves_home if side=='home' else saves_away,
                                                   playername_on_sheet, player_slapid_on_sheet, playerstats['Score'],
                                                   playerstats['Goals'], playerstats['Assists'], playerstats['Shots'],
                                                   playerstats['Saves'], playerstats['Passes'], playerstats['Blocks'],
                                                   playerstats['Post Hits'], playerstats['Takeaways'], playerstats['Turnovers'],
                                                   playerstats['Primary Assists'], playerstats['Secondary Assists'],
                                                   playerstats['Possession Time'], playerstats['Faceoffs Won'],
                                                   playerstats['Faceoffs Lost'], playerstats['Wins'], playerstats['Losses'],
                                                   playerstats['Ties'], playerstats['OT Wins'], playerstats['OT Losses'],
                                                   playerstats['Game Winning Goals'], playerstats['Shutout Periods'],
                                                   playerstats['Shutout Periods Against'], playerstats['Goals For'],
                                                   playerstats['Goals Against'], playerstats['Periods Played'], away_team if side == 'home' else home_team, 'OT' if away_ot_losses > 0 or home_ot_losses > 0 else '', '', '', '', '']], columns=match_df.columns)
                    temp_df = temp_df.append(player_df) #add the player dataframe as a row to the temporary dataframe

            #once we have looped through all the players in this "side", convert scores to numeric and sort
            temp_df["Score"] = pd.to_numeric(temp_df["Score"], errors='ignore')
            temp_df=temp_df.sort_values('Score', ascending=False)
            return match_df.append(temp_df)

        def hidden_df_maker(side):
            '''This makes the hidden line that is needed for stats retrieval'''
            hidden_list = ['', '', match_code, 'WIN' if side == winner else 'LOSS', home_team if side == 'home' else away_team,
                           goals_home if side == 'home' else goals_away,
                           goals_away if side == 'home' else goals_home, shots_home if side == 'home' else shots_away,
                           shots_away if side == 'home' else shots_home,
                           saves_home if side == 'home' else saves_away]
            append_list = ['' for i in range(len(match_df.columns)-len(hidden_list))]
            append_list[28] = away_team if side =='home' else home_team
            append_list[29] = 'OT' if away_ot_losses > 0 or home_ot_losses > 0 else ''
            hidden_list.extend(append_list)
            hidden_df = pd.DataFrame([hidden_list], columns=match_df.columns)
            return hidden_df

        #add the column
        match_df = match_df.append(column_df)
        #then add the home team
        match_df = team_checker(log, 'home', match_df)
        for i in range(6 - playercount_home):
            #we make the hidden df for every empty spot - depends on playercount
            match_df = match_df.append(hidden_df_maker('home'))

        #columns again for away team
        match_df = match_df.append(column_df)
        #add away team stats
        match_df = team_checker(log, 'away', match_df)
        for i in range(6 - playercount_away):
            match_df = match_df.append(hidden_df_maker('away'))

        #add the notes line
        notes_list = ['' for _ in range(len(match_df.columns))]
        notes_list[2] = 'notes:'
        notes_list[3] = 'unverified'
        notes_df = pd.DataFrame([notes_list], columns=match_df.columns)
        match_df = match_df.append(notes_df)
        match_df = match_df.append(blank_df)
        #again rename trick to be able to have duplicate column names
        match_df.rename(columns={'blank1':'', 'blank2':'', 'TScore': 'Score', 'TGoals Against':'Goals Against','TShots': 'Shots', 'TSaves': 'Saves', 'loss or win':'', 'some_shit':'', 'match_code':''}, inplace=True)
        return match_df

def rename_json_team_date(filename):
    '''This function renames the JSON files to the appropriate format
    Requirement is that JSON files are in their original format of YYYY-MM-DD-HH-MM-SS.json'''
    log = import_json(filename)
    recognize_team(log)
    datetime_object = datetime.strptime(filename[:-5], '%Y-%m-%d-%H-%M-%S')
    new_name = 'S{} {} {} {} - {} {} {}.json'.format(season, format, div, recognize_team(log)[1], recognize_team(log)[0], datetime_object.date(), log['current_period'])

    #failsafe for 2 matches finishing the exact same second (lol)
    if os.path.exists(new_name):
        for i in range(2, 10):
            new_name = 'S{} {} {} {} - {} {} ({}) {}.json'.format(season, format, div, recognize_team(log)[1], recognize_team(log)[0], datetime_object.date(), i, log['current_period'])
            if not os.path.exists(new_name):
                break
    os.rename(filename, new_name)


def export_to_excel():
    '''This function gets all the jsons, renames them, converts them to match logs
    Then adds them all together, and exports it to an excel feel'''
    print('Getting all json files...')
    json_filenames = [pos_json for pos_json in os.listdir('.') if pos_json.endswith('.json')]
    result = pd.DataFrame()
    print('Converting json files to match reports...')
    for json_filename in json_filenames:
        json_file = import_json(json_filename)
        rename_json_team_date(json_filename)
        match_df = log_to_match(json_file)
        result = pd.concat([match_df, result])

    print('Exporting match reports to excel...')
    ##get current directory
    cwd = os.getcwd()
    stats_name = f'{cwd}\\S{season} {format} {div} W{week} stats.xlsx'
    #here we export toe xcel, putting index and header to false cause we dont want those
    result.to_excel(r''+stats_name, sheet_name='stats', index=False, header=False)
    print('Done!')



def apply_formatting():
    cwd = os.getcwd()
    ws = openpyxl.load_workbook(f'{cwd}\\S{season} {format} {div} W{week} stats.xlsx')
    wb = ws['stats']

    red_fill = PatternFill(patternType='solid', fgColor='dd3838')
    blue_fill = PatternFill(patternType='solid', fgColor='3d85c6')
    gray_fill = PatternFill(patternType='solid', fgColor='b7b7b7')
    red_font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='dd3838')
    blue_font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='3d85c6')
    gray_font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='b7b7b7')
    black_font = Font(name='Arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='000000')

    white_rows = [i for i in range(1, 999, 19)]
    gray_rows = [i for i in range(2, 999, 1) if i not in white_rows]
    start_red = [2,3,4,5,6,7,8,9]
    start_red_font = [4,5,6,7,8]
    start_blue = [9,10,11,12,13,14,15]
    start_blue_font = [11,12,13,14,15]
    start_black_font = [3,10,17]

    red_rows = [(i+j) for j in range(2,10000,19) for i in start_red]
    blue_rows = [(i+j) for j in range(2,10000,19) for i in start_blue]
    red_font_rows = [(i+j) for j in range(2,10000,19) for i in start_red_font]
    blue_font_rows = [(i+j) for j in range(2,10000,19) for i in start_blue_font]
    black_font_rows = [(i+j) for j in range(2,10000,19) for i in start_black_font]
    gray_match_color = [(i+j) for j in range (1,10000,19) for i in range(2,14)]

    for row in wb.iter_rows(min_row=1, max_row=wb.max_row, min_col=1, max_col=wb.max_column):
        for cell in row:
            if cell.row in white_rows:
                continue
            if cell.column in [1,2]:
                continue
            if cell.row in gray_rows:
                cell.fill = gray_fill
                if cell.row in black_font_rows:
                    cell.font = black_font
                elif cell.row in gray_match_color and cell.column == 3:
                    cell.font = gray_font
            if cell.row in red_rows:
                cell.fill = red_fill
                if cell.column in [3,4] or cell.column > 38:
                    cell.fill = gray_fill
                    cell.font = gray_font
                    if cell.column > 38:
                        cell.font = black_font
            if cell.row in blue_rows:
                cell.fill = blue_fill
                if cell.column in [3,4] or cell.column > 38:
                    cell.fill = gray_fill
                    cell.font = gray_font
                    if cell.column > 38:
                        cell.font = black_font
            if cell.row in red_font_rows:
                if cell.column in [5,6,7,8,9,10]:
                    cell.font = red_font
            if cell.row in blue_font_rows:
                if cell.column in [5,6,7,8,9,10]:
                    cell.font = blue_font
            if cell.row in black_font_rows:
                if cell.column == 4:
                    cell.font = black_font

    filename_c = f'S{season} {format} {div} W{week} stats color coded.xlsx'
    ws.save(filename_c)

#call the export to excel function
export_to_excel()

#an excel file now exists that we can apply proper formatting to:
apply_formatting()





